import json
import requests
import sys
import os
import csv
import openpyxl
from openpyxl import load_workbook
from pathlib import Path
import re


parent_dir = Path(__file__).resolve().parent
BDC_provider = parent_dir  / ".." / "Data_set" / "FCC_providers" / "BDC.xlsx"
folder_path = parent_dir  / ".." / "Data_set" / "FCC_providers" 
output_path = parent_dir  / ".." / "FCC_Provider" / "FCC_providers.txt"
output_path2 = parent_dir  / ".." / "outputs" / "FCC_providers.txt"

providers = []
fingers = []

class Provider:
    def __init__(self, frn, provider_id, brand, br):
        self.frn = frn
        self.provider_id = provider_id
        self.brand = brand
        self.br = br
        self.if_provider_info = 0 #if we could use frn and provider id to find provider in provideridtable_2024_06_28_1204, this bit is 1, else 0
    
    def get_frn(self):
        return self.frn
    
    def get_provider_id(self):
        return self.provider_id
    
    def get_brand(self):
        return self.brand
    
    def get_br(self):
        return self.br

    def generate_finger(self):
        return self.frn + "||" + self.provider_id + "||" + self.brand + "||" + self.br

    def generate_f(self):
        return self.frn + "||" + self.provider_id + "||" + self.brand      

    def get_provider_info(self):
        #print("1111")

        workbook = load_workbook(BDC_provider)
        sheet = workbook['provideridtable_2024_06_28_1204']
        for row in sheet.iter_rows(values_only=True):
            #print("the type of provider id is: " + str(type(row[4])))
            #print("the type of frn is: " + str(type(row[3])))            
            if row[4] == int(self.provider_id) and row[3] == self.frn:
                #print("find the provider") 
                #print(row)
                self.if_provider_info = 1
                self.provider_name = row[0]
                self.holding_company = row[1]
                self.operation_type = row[2]
                return 

    def modify_br(self, new_br):
        self.br = new_br

    def print_provider(self):
        data = self.__dict__.copy()  
        del data['if_provider_info']  
        if self.if_provider_info == 0:
            del data['provider_name']
            del data['provider_company']
            del data['operation_type']
        return data


def build_arg(street, city, state):
    paramas = {
        "street": street,
        "city": city,
        "state": state,
        "benchmark": "Public_AR_Current",
        "vintage": "Current_Current",
        "format": "json"
    }
    return paramas


def frn_search(folder_name, geoid):
    if not os.path.isdir(folder_name):
        print(f"The folder '{folder_name}' does not exist.")
        return
    for root, _, files in os.walk(folder_name):
        for file in files:
            if file.endswith('.csv'):
                file_path = os.path.join(root, file)
                with open(file_path, mode='r', encoding='utf-8') as csv_file:
                    reader = csv.DictReader(csv_file)
                    for row in reader:
                        #print(row.get('block_geoid'))
                        if row.get('block_geoid') == geoid:
                            frn = row.get('frn')
                            p_id = row.get('provider_id')
                            b_name = row.get('brand_name')
                            br = row.get('business_residential_code')
                            p = Provider(frn, p_id, b_name, br)
                            if p.generate_finger() not in fingers:
                                print(f"---Match found in file: {file_path}")
                                print("---The frn code of provider is: " + frn)
                                print("---The provider_id of provider is: " + p_id)
                                print("---The brand name of the provider is: " + b_name)
                                if  br == "R":
                                    print("---This is Residential-only location")
                                elif br == "B":
                                    print("---This is Business-only location")
                                elif br == "X":
                                    print("---This is Business and Residential location")
                                providers.append(p)
                                fingers.append(p.generate_finger())



def FIPS_lookup(street, city, state, zip):

    #print(f"the state is '{state}'")
    URL = 'https://geocoding.geo.census.gov/geocoder/geographies/address'
    paramas = build_arg(street, city, state)
    response = requests.get(URL, paramas)
    data = response.json()
    ddd = data['result']['addressMatches']
    for i in range(len(data['result']['addressMatches'])):
        json_zip = data['result']['addressMatches'][i]["addressComponents"]["zip"]
        if json_zip == zip:
            target_geoid = data['result']['addressMatches'][i]["geographies"]["2020 Census Blocks"][0]["GEOID"]
            return target_geoid
        else:
            #print("False: " + street + " " + json_zip + " " + data['result']['addressMatches'][i]["geographies"]["2020 Census Blocks"][0]["GEOID"])
            return "0"
    return "0"








if len(sys.argv) > 1:
    address = sys.argv[1]  # Get the first argument (aaaaa)
    add = address.split(', ')
    street = add[0]
    city = add[1]
    state = add[2]
    zip = add[3]
    match = re.search(r'\d+', street)
    target_geoid = FIPS_lookup(street,city,state,zip)
    print(f"The FIPS code for {street} is: {target_geoid}")
    if target_geoid == "0":
        match = re.search(r'\d+', street)
        street_num = int(match.group())
        for i in range(50):
            new_add = street_num + i*2
            updated_address = re.sub(r'\d+', str(new_add), street, 1)
            geoid = FIPS_lookup(updated_address,city,state,zip)
            if geoid != "0":
                target_geoid = geoid
                print(f"The FIPS code for {updated_address} is: {target_geoid}")
                break
            new_add = street_num - i*2
            updated_address = re.sub(r'\d+', str(new_add), street, 1)
            geoid = FIPS_lookup(updated_address,city,state,zip)
            if geoid != "0":
                target_geoid = geoid
                print(f"The FIPS code for {updated_address} is: {target_geoid}")
                break
    print(f"current geoid is {target_geoid}")
    if target_geoid == "0":
        match = re.search(r'\d+', street)
        street_num = int(match.group()) + 1
        for i in range(50):
            new_add = street_num + i*2
            updated_address = re.sub(r'\d+', str(new_add), street, 1)
            geoid = FIPS_lookup(updated_address,city,state,zip)
            if geoid != "0":
                target_geoid = geoid
                print(f"The FIPS code for {updated_address} is: {target_geoid}")
                break
            new_add = street_num - i*2
            updated_address = re.sub(r'\d+', str(new_add), street, 1)
            geoid = FIPS_lookup(updated_address,city,state,zip)
            if geoid != "0":
                target_geoid = geoid
                print(f"The FIPS code for {updated_address} is: {target_geoid}")
                break        
    '''
    if match:
        street_num = int(match.group())
        for i in range(50):
            new_add = street_num + i
            updated_address = re.sub(r'\d+', str(new_add), street, 1)
            geoid = FIPS_lookup(updated_address,city,state,zip)
            if geoid == target_geoid and target_geoid != "0":
                print(updated_address)
            new_add = street_num - i
            updated_address = re.sub(r'\d+', str(new_add), street, 1)
            geoid = FIPS_lookup(updated_address,city,state,zip)
            if geoid == target_geoid and target_geoid != "0":
                print(updated_address)
    else:
        geoid = FIPS_lookup(street,city,state,zip)
        print(f"The fips code for {updated_address} is {geoid}")
    '''
    print("+++++++++++++++++++++++++++++++++++")
    folder_path = folder_path / state.strip()
    frn_search(folder_path, target_geoid)
    print("*********************************")
    for i in range(len(providers)):
        j = i + 1
        while j < len(providers):
            if providers[i].generate_f() == providers[j].generate_f():
                if providers[i].get_br() != providers[j].get_br():
                    providers[i].modify_br('X')
                providers.pop(j)
            else:
                j = j + 1
    output_list = []
    for pro in providers:
        pro.get_provider_info()
        output_list.append(pro.print_provider())
    json_output = json.dumps(output_list)
    print(json_output)
    with open(output_path, "w") as file:
        file.write(json_output)
    with open(output_path2, "w") as file:
        file.write(json_output)
else:
    print("Please provide the physical address of the library")

